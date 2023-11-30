import re
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime


def validCell(cell: str) -> int:
    if cell != 'No value':
        if cell.find("K") > 0:
            string_num = extract_numbers(cell)
            number = float(string_num) * 1000
            number = int(number)
        else:
            string_num = extract_numbers(cell)
            number = int(float(string_num)) 
    return number


def validColumn(wb: Workbook) -> int:
    sheet = wb.active
    last_col = 1
    for col_idx in range(3,27):
        column_letter = get_column_letter(col_idx)
        cell = sheet[f'{column_letter}{2}']
        if (type(cell.value) == str) and (cell.value != "No value"):
            last_col = col_idx
            for i in range(1,30):
                if sheet[f'{column_letter}{i}'].value != "No value":
                    sheet[f'{column_letter}{i}'].value = validCell(sheet[f'{column_letter}{i}'].value)
            break
    return last_col

                
def set_formula_and_time(wb: Workbook, last_col: int):
    sheet = wb.active
    last_col_letter = get_column_letter(last_col)
    penultimate = get_column_letter(last_col - 1)
    for i in range(1, 30):
        if sheet[f'{penultimate}{i}'].value != 'No value':
            sheet[f'B{i}'] = f'={last_col_letter}{i} - {penultimate}{i}'
        else:
            sheet[f'B{i}'].value = 'No value'
    sheet['B30'] = datetime.today().date()
    sheet['B31'] = str(datetime.time(datetime.today()))


def count_participants(wb: Workbook, last_col: int):
    sheet = wb.active
    sum = 0
    col_letter = get_column_letter(last_col)
    for j in range(1, 30):
        if type(sheet[f'{col_letter}{j}'].value) is int:
            sum += sheet[f'{col_letter}{j}'].value
    sheet[f'{last_col}{32}'].value = sum


def extract_numbers(input_string):
    numbers = re.findall(r'\d+\.\d+|\d+', input_string)
    result = ''.join(numbers)
    return result
