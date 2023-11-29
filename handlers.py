import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException
import os
from datetime import datetime

def createWorkbook(file_name: str):
    if os.path.exists(file_name):
        file_path = 'stats.xlsx'
        wb = load_workbook(filename=file_path)
    else:
        wb = Workbook()
    return wb

def work(links: dict, wb: Workbook):
    driver = webdriver.Chrome()
    sheet = wb.active
    col = 2
    for value in links.values():
        driver.get(f'{value[0]}')
        
        block = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CLASS_NAME, 'list-container')))
       
        start_xlsx = (int)(value[1])
        loop_end = (int)(value[2])
        value_start = sheet[f'A{start_xlsx}'].value
        if value_start is None:
            for i in range(1, loop_end - start_xlsx + 1):
                xpath_name = "//*[@id='ga-campaign-collection-page']/div/div[2]/div[2]/div[" + (str)(i) + "]/a/div/div[2]/div[1]/h1"
                event_name = block.find_element(By.XPATH, xpath_name)
                sheet[f'A{start_xlsx + i - 1}'] = event_name.text
                
                column_letter = get_column_letter(2)
                
                sheet[f'{column_letter}{start_xlsx + i - 1}'] = 0
                sheet['B30'] = datetime.today().date()
        else:
            for col_idx in range(3,27):
                column_letter = get_column_letter(col_idx)
                cell = sheet[f'{column_letter}{start_xlsx}']
                
                if cell.value is None:
                    for i in range(1, loop_end - start_xlsx + 1):
                        xpath_count = "//*[@id='ga-campaign-collection-page']/div/div[2]/div[2]/div[" + (str)(i) + "]/a/div/div[2]/div[1]/div[2]/div"
                        cell = sheet[f'{column_letter}{start_xlsx + i - 1}']
                        try:
                            participants = block.find_element(By.XPATH, xpath_count)
                            sheet[f'{column_letter}{start_xlsx + i - 1}'] = participants.text
                        except NoSuchElementException:
                            sheet[f'{column_letter}{start_xlsx + i - 1}'] = "No value"
                    sheet[f'{column_letter}30'] = datetime.today().date()
                    break
    driver.quit()

def run():
    links = {
    '1': ('https://galxe.com/Galxe/campaign/GCmoQtUMDL', '1', '12'),
    '2': ('https://galxe.com/Galxe/campaign/GCfLytU9Qf', '12', '27'),
    '3': ('https://galxe.com/Galxe/campaign/GCih6tUwoE', '27', '30')
    }
    file_name = 'stats.xlsx'
    wb = createWorkbook(file_name)
    work(links, wb)
    wb.save(filename='stats.xlsx')